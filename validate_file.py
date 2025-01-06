from PIL import Image
import os
import re
import sys
import csv
import warnings
import structlog
import logging
from datetime import datetime, timedelta
from time import time
import home_automation_common
import argparse

""" 
    This Python script validates files of various types (e.g., images, documents, videos) in a specified directory. 
    It uses command-line arguments for input and incorporates structured logging.
"""

warnings.simplefilter("ignore", UserWarning)


# Define logical groupings for file types - also used in collector.py
FILE_TYPE_GROUPS = {
    "image": [r".*\.(jpg|jpeg|png|gif|bmp|tiff)$"],
    "document": [r".*\.(docx)$"],
    "video": [r".*\.(mp4|avi|mkv|mov|flv|wmv|webm)$"],
    "audio": [r".*\.(mp3|wav|aac|flac|m4a|ogg)$"],
    "excel": [r".*\.(xlsx)$"],
    "pdf": [r".*\.(pdf)$"],
}


def _get_arguments():
    """
    Parses command-line arguments for directory and file type.
    Returns:
        argparse.Namespace: A namespace object containing the parsed arguments.
            - directory (str): Path to the directory to process. Defaults to 'F:\\'.
            - filetype (str): Type of files to process (e.g., 'image', 'video', etc.). Defaults to 'image'.
    """
    parser = argparse.ArgumentParser(
        description="Process a directory and file type for file operations."
    )

    # Add named arguments
    parser.add_argument(
        "--directory",
        "-d",
        type=str,
        default="F:\\",
        help="Path to the directory to process. Defaults to 'F:\\'.",
    )
    parser.add_argument(
        "--filetype",
        "-f",
        type=str,
        default="image",
        help="Type of files to process (e.g., 'image', 'video', etc.). Defaults to 'image'.",
    )

    # Parse the arguments
    args = parser.parse_args()

    # Return arguments as a namespace object
    return args


def _validate_document(file_path):
    """
    Validates a Word document file.
    This function attempts to open a Word document file using the `python-docx` library.
    It checks if the document contains any paragraphs to determine its validity.
    Args:
        file_path (str): The path to the Word document file to be validated.
    Returns:
        tuple: A tuple containing a boolean and a string message.
               - The boolean indicates whether the document is valid (True) or not (False).
               - The string message provides additional information about the validation result.
    """
    from docx import Document

    try:
        # Try to open the document
        doc = Document(file_path)

        # Check if the document contains paragraphs
        if len(doc.paragraphs) > 0:
            return True, "Valid document file."
        else:
            return False, f"Document {file_path} is empty."

    except Exception as e:
        return False, f"Document {file_path} is not valid."


def _validate_audio(file_path):
    """
    Validates if the given file path points to a valid audio file.
    This function attempts to load the audio file using the pydub library.
    If the file is successfully loaded, it is considered a valid audio file.
    Otherwise, it is considered invalid.
    Args:
        file_path (str): The path to the audio file to be validated.
    Returns:
        tuple: A tuple containing a boolean and a string.
               The boolean indicates whether the file is a valid audio file.
               The string provides a message with the validation result.
    """
    from pydub import AudioSegment

    try:
        # Attempt to load the audio file
        audio = AudioSegment.from_file(file_path)
        return True, f"Valid audio file: {file_path}"
    except Exception as e:
        return False, f"Invalid audio file: {file_path} (Error: {e})"


def _validate_excel(file_path):
    """
    Validates an Excel (.xlsx) file by attempting to open it and checking for the presence of sheets.
    Args:
        file_path (str): The path to the Excel file to be validated.
    Returns:
        tuple: A tuple containing a boolean and a string.
            - bool: True if the file is a valid .xlsx file with at least one sheet, False otherwise.
            - str: A message indicating the result of the validation.
    """
    from openpyxl import load_workbook

    try:
        # Try to open the workbook
        workbook = load_workbook(file_path, read_only=True)
        # Check if the workbook contains any sheets
        if not workbook.sheetnames:
            return False, f"Invalid .xlsx file: {file_path} (No sheets found)."
        else:
            return True, "Valid .xlsx file."
    except Exception as e:
        return False, f"Invalid .xlsx file: {file_path} (Error: {e})."


def _validate_pdf(file_path):
    """
    Validates a PDF file.
    This function checks if the provided file path points to a valid PDF file.
    It uses the PyPDF2 library to read the PDF and determine if it is valid,
    has pages, and whether it is encrypted.
    Args:
        file_path (str): The path to the PDF file to be validated.
    Returns:
        tuple: A tuple containing a boolean and a string message.
               The boolean indicates whether the PDF is valid.
               The string provides additional information about the validation result.
               - (True, "File is encrypted") if the PDF is valid and encrypted.
               - (True, "Valid pdf file.") if the PDF is valid and not encrypted.
               - (False, "Invalid PDF: {file_path} (No pages found)") if the PDF has no pages.
               - (False, "Invalid PDF: {file_path} (Error: {e})") if an error occurred during validation.
    """
    from contextlib import redirect_stderr
    import io

    try:
        # Importing PyPDF2 inside the function
        from PyPDF2 import PdfReader

        with io.StringIO() as err_buffer, redirect_stderr(err_buffer):
            reader = PdfReader(file_path)

            if reader.pages:
                if reader.is_encrypted:
                    return True, "File is encrypted"
                else:
                    return True, "Valid pdf file."
            else:
                return False, f"Invalid PDF: {file_path} (No pages found)"

    except Exception as e:
        return False, f"Invalid PDF: {file_path} (Error: {e})"


def _validate_video(file_path):
    """
    Validates a video file using FFprobe.
    This function runs the FFprobe command to check the duration of the video file,
    which serves as an indicator of its validity. If the command executes successfully,
    the video file is considered valid.
    Args:
        file_path (str): The path to the video file to be validated.
    Returns:
        tuple: A tuple containing a boolean and a string message.
               - True and "Valid video file." if the video file is valid.
               - False and an error message if the video file is not valid.
    """
    import subprocess

    try:
        # Run FFprobe command
        command = [
            "ffprobe",
            "-v",
            "error",  # Only show errors
            "-show_entries",
            "format=duration",  # Check for duration (validity indicator)
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            file_path,
        ]
        subprocess.check_output(command, stderr=subprocess.STDOUT)
        return True, "Valid video file."
    except subprocess.CalledProcessError as e:
        return False, f"Video file {file_path} is not a valid video file."


def _validate_image(file_path):
    """
    Validates whether the given file path points to a valid image file.

    Args:
        file_path (str): The path to the image file to be validated.

    Returns:
        tuple: A tuple containing a boolean and a string message.
               The boolean is True if the file is a valid image, False otherwise.
               The string message provides additional information about the validation result.
    """
    try:
        with Image.open(file_path) as img:
            img.verify()  # Verify that it is a valid image
        return True, "Valid image file."  # the file is not invalid
    except Exception as e:
        return False, f"Invalid image file: {e}"


def _get_total_file_count(directory, exclusions):
    """
    Calculate the total number of files in a directory, excluding specified files and directories.
    Args:
        directory (str): The root directory to start the file count.
        exclusions (list): A list of file and directory paths to exclude from the count.
    Returns:
        int: The total number of files in the directory, excluding the specified exclusions.
    """
    file_count = 0

    for root, dirs, files in os.walk(directory):
        # Exclude directories
        dirs[:] = [d for d in dirs if os.path.join(root, d) not in exclusions]

        # Count files, excluding those in the exclusions
        for file in files:
            file_path = os.path.join(root, file)
            if file_path not in exclusions:
                file_count += 1

    return file_count


def _normalize_path(path, directory=None):
    """
    Normalize the given file path.

    If a directory is provided, it will be prepended to the path before normalization.

    Args:
        path (str): The file path to normalize.
        directory (str, optional): The directory to prepend to the path. Defaults to None.

    Returns:
        str: The normalized file path.
    """
    if directory:
        path = f"{directory}{path}"
    return os.path.normpath(path)


def _get_file_count_for_type(directory, compiled_pattern, exclusions):
    """
    Count the number of files in a directory and its subdirectories that match a given pattern,
    excluding specified files and directories.
    Args:
        directory (str): The root directory to start the search.
        compiled_pattern (re.Pattern): The compiled regular expression pattern to match file names.
        exclusions (list): A list of file and directory paths to exclude from the count.
    Returns:
        int: The count of files that match the pattern and are not in the exclusions list.
    """
    file_count = 0

    # Walk through the directory and subdirectories
    for root, dirs, files in os.walk(directory):
        # Exclude directories
        dirs[:] = [d for d in dirs if os.path.join(root, d) not in exclusions]

        # Count matching files, excluding those in the exclusions
        for file in files:
            file_path = os.path.join(root, file)
            if file_path not in exclusions and compiled_pattern.match(file):
                file_count += 1

    return file_count


def _write_summary_file(
    file_type_or_group, total_files, file_count, error_count, start_time, end_time
):
    """
    Writes a summary of the file validation process to a CSV file.
    Args:
        file_type_or_group (str): The type or group of files being validated.
        total_files (int): The total number of files processed.
        file_count (int): The number of files that matched the criteria.
        error_count (int): The number of files that encountered errors.
        start_time (datetime): The start time of the validation process.
        end_time (datetime): The end time of the validation process.
    Returns:
        None
    """

    summary_headers = [
        "date",
        "file_type",
        "total_file_count",
        "matching_file_count",
        "error_count",
        "start_time",
        "end_time",
        "duration",
    ]
    summary_output_file = "validation_summary_output.csv"

    duration = home_automation_common.duration_from_times(end_time, start_time)

    summary_output_data = [
        today,
        file_type_or_group,
        total_files,
        file_count,
        error_count,
        start_time,
        end_time,
        duration,
    ]
    # headers = ["Date", "Description", "Details"]

    _append_to_csv(summary_output_file, summary_output_data, summary_headers)


def _append_to_csv(file_path, data, headers=None):
    """
    Appends data to a CSV file. If the file does not exist, it creates a new one and optionally writes headers.
    Args:
        file_path (str): The path to the CSV file.
        data (list): The data to append to the CSV file.
        headers (list, optional): The headers to write if the file is newly created. Defaults to None.
    Returns:
        None
    """

    output_file = home_automation_common.get_full_filename("output", file_path)

    file_exists = os.path.exists(output_file)  # Check if the file already exists

    with open(output_file, mode="a", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)

        # Write headers if the file is newly created
        if not file_exists and headers:
            writer.writerow(headers)

        # Append the data
        writer.writerow(data)


def validate_files_by_type(start_folder, file_type_or_group):
    """
    Validates files in a given directory based on their type or group.
    Args:
        start_folder (str): The directory to search for files.
        file_type_or_group (str or list): The file type or group to validate.
                                          Can be a string representing a single file type or a group,
                                          or a list of file types.
    Raises:
        ValueError: If the file_type_or_group is not a string or list, or if the start_folder does not exist.
    Logs:
        Various informational and error messages during the validation process.
    The function performs the following steps:
        1. Logs the start of the search.
        2. Resolves regex patterns based on the file type or group.
        3. Checks if the start folder exists.
        4. Compiles all patterns into a single regex for efficiency.
        5. Loads exclusions from an exclusion file.
        6. Logs the beginning of the file type search.
        7. Counts the total files and matching files.
        8. Logs the number of matching files found.
        9. Writes headers to an output CSV file.
        10. Walks through the directory and subdirectories to validate files.
        11. Logs any validation exceptions found.
        12. Removes the output file if no errors are found.
        13. Writes a summary file.
        14. Logs the completion of the validation.
    Returns:
        None
    """
    from tqdm import tqdm

    start_time = datetime.now().time()

    logger = structlog.get_logger()

    logger.info(
        "Searching for file types.",
        module="validate_file.validate_files_by_type",
        message="Directory to be searched for file types.",
        folder=start_folder,
        type=file_type_or_group,
    )

    # Resolve regex patterns based on group or specific type
    if isinstance(file_type_or_group, str):
        patterns = FILE_TYPE_GROUPS.get(
            file_type_or_group.lower(), [file_type_or_group]
        )
    elif isinstance(file_type_or_group, list):
        patterns = file_type_or_group
    else:
        logger.error(
            "Invalid file type argument.",
            module="validate_file.validate_files_by_type",
            message="file_type_or_group must be a string or a list",
        )
        raise ValueError("file_type_or_group must be a string or a list")

    if not os.path.exists(start_folder):
        logger.error(
            "Folder does not exist",
            module="validate_file.validate_files_by_type",
            message=f"Folder {start_folder} does not exist. Retry.",
        )
        raise ValueError("Invalid folder was specified. It does not exist.")

    # Compile all patterns into a single regex for efficiency
    combined_pattern = re.compile("|".join(patterns), re.IGNORECASE)

    exclusion_file = f"{file_type_or_group}_exclusions.txt"

    # Load exclusions
    exclusions = set()
    if exclusion_file:
        try:
            with open(exclusion_file, "r", encoding="utf-8") as f:
                # exclusions = set(line.strip() for line in f if line.strip())
                exclusions = set(
                    _normalize_path(line.strip(), start_folder)
                    for line in f
                    if line.strip()
                )
        except FileNotFoundError:
            logger.info(
                "No exclusions found.",
                module="validate_file.validate_files_by_type",
                message=f"Exclusion file {exclusion_file} not found. Continuing without exclusions.",
            )

    file_count = 0
    error_count = 0
    logger.info(
        "File type search beginning.",
        module="validate_file.validate_files_by_type",
        message=f"Looking for {file_type_or_group} files in directory {start_folder}.",
    )

    total_files = _get_total_file_count(start_folder, exclusions)
    matching_files = _get_file_count_for_type(
        start_folder, combined_pattern, exclusions
    )

    logger.info(
        "Matching files found.",
        module="validate_file.validate_files_by_type",
        message=f"Found a total of {matching_files} found. Validating files now.",
    )

    headers = ["file_name", "error_message"]
    output_file = f"{datetime.now().date()}-error-output-{file_type_or_group}.csv"

    output_file = home_automation_common.get_full_filename("output", output_file)

    with open(output_file, mode="w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)  # Write the headers

        with tqdm(
            total=matching_files, desc="Processing matching files", unit="file"
        ) as pbar:  # Walk through the directory and subdirectories
            for root, dirs, files in os.walk(start_folder):
                # Skip excluded directories
                dirs[:] = [d for d in dirs if os.path.join(root, d) not in exclusions]

                for filename in files:
                    file_path = os.path.join(root, filename)

                    # Skip excluded files
                    if file_path in exclusions:
                        continue

                    if combined_pattern.match(filename):
                        file_count += 1
                        pbar.set_description(f"Processing: {filename}")
                        file_path = r"\\?\\" + os.path.abspath(file_path)
                        # Validate file based on type
                        match file_type_or_group:
                            case "image":
                                is_valid, error_message = _validate_image(file_path)
                            case "pdf":
                                is_valid, error_message = _validate_pdf(file_path)
                            case "video":
                                is_valid, error_message = _validate_video(file_path)
                            case "excel":
                                is_valid, error_message = _validate_excel(file_path)
                            case "audio":
                                is_valid, error_message = _validate_audio(file_path)
                            case "document":
                                is_valid, error_message = _validate_document(file_path)
                            case _:
                                logger.error(
                                    "Incorrect file type found.",
                                    module="validate_file.validate_files_by_type",
                                    message="An undefined filetype is found",
                                )
                                is_valid, error_message = False, "Undefined filetype"

                        pbar.update(1)
                        if not is_valid:
                            error_count += 1
                            writer.writerow([file_path, error_message])
                            logger.info(
                                "File validation exception found.",
                                module="validate_file.validate_files_by_type",
                                message=error_message,
                                file=file_path,
                            )

    if error_count == 0:
        os.remove(output_file)

    end_time = datetime.now().time()

    _write_summary_file(
        file_type_or_group, total_files, file_count, error_count, start_time, end_time
    )

    logger.info(
        "Validation completed.",
        module="validate_file.validate_files_by_type",
        message="Analysis complete.",
        total_files=total_files,
        matching_files=matching_files,
        error_count=error_count,
    )


if __name__ == "__main__":

    today = datetime.now().date()

    log_file = f"{today}_validate_file_log.txt"

    log_file = home_automation_common.get_full_filename("log", log_file)

    home_automation_common.configure_logging(log_file)

    logger = structlog.get_logger()

    args = _get_arguments()

    # Validate files
    validate_files_by_type(args.directory, args.filetype)
