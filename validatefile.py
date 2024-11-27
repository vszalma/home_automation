from PIL import Image
import os
import re
import sys
import csv

# Define logical groupings for file types
FILE_TYPE_GROUPS = {
    "image": [r".*\.(jpg|jpeg|png|gif|bmp|tiff)$"],
    "document": [r".*\.(docx)$"],
    "video": [r".*\.(mp4|avi|mkv|mov|flv|wmv|webm)$"],
    "audio": [r".*\.(mp3|wav|aac|flac|ogg)$"],
    "excel": [r".*\.(xlsx)$"],
    "pdf": [r".*\.(pdf)$"],
}


def list_files_by_regex(start_folder, file_type_or_group):
    """
    List files in the given folder and subfolders filtered by a specific file type or logical group using regex.

    Parameters:
        start_folder (str): The root folder to start searching from.
        file_type_or_group (str or list): The file type(s) to filter by (e.g., "*.pdf")
                                          or a logical group (e.g., "images").

    Returns:
        list: A list of matching file paths.
    """
    from datetime import datetime
    from tqdm import tqdm
    
    # Resolve regex patterns based on group or specific type
    if isinstance(file_type_or_group, str):
        patterns = FILE_TYPE_GROUPS.get(
            file_type_or_group.lower(), [file_type_or_group]
        )
    elif isinstance(file_type_or_group, list):
        patterns = file_type_or_group
    else:
        raise ValueError("file_type_or_group must be a string or a list")

    # Compile all patterns into a single regex for efficiency
    combined_pattern = re.compile("|".join(patterns), re.IGNORECASE)

    file_count = 0
    error_count = 0
    print(f"Looking for {file_type_or_group} files in directory {start_folder}.")

    total_files = get_total_file_count(start_folder)
    matching_files = count_files_with_pattern(start_folder, combined_pattern)

    print(f"Found a total of {matching_files} found. Validating files now.")
    
    headers = ["file_name", "error_message"]
    output_file = f"{datetime.today().strftime('%Y-%m-%d')}-error-output-{file_type_or_group}.csv"

    with open(output_file, mode="w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(["file_name", "error_message"])  # Write the headers
 
        with tqdm(total=matching_files, desc="Processing matching files", unit="file") as pbar:       # Walk through the directory and subdirectories
            for root, dirs, files in os.walk(start_folder):
                for filename in files:
                    if combined_pattern.match(filename):
                        file_count += 1
                        invalid_file = False
                        file_name = f"{root}\\{filename}"
                        # matching_files.append(os.path.join(root, filename))
                        match file_type_or_group:
                            case "image":
                                is_valid, error_message = validate_image(f"{file_name}")
                            case "pdf":
                                is_valid, error_message = validate_pdf(f"{file_name}")
                            case "video":
                                is_valid, error_message = validate_video(f"{file_name}")
                            case "excel":
                                is_valid, error_message = validate_excel(f"{file_name}")
                            case "document":
                                is_valid, error_message = validate_document(f"{file_name}")
                            case _:
                                print("An undefined filetype is found")
                        pbar.update(1)
                        if not is_valid:
                            error_count += 1
                            # pass filename {root}\\{filename} and error_message to method to write file
                            writer.writerow([file_name, error_message])
                        
    if error_count == 0:
        os.remove(output_file)

    print(f"Total files read: {file_count}")
    print(f"Total errors found: {error_count}")

def validate_document(file_path):
    from docx import Document

    try:
        # Try to open the document
        doc = Document(file_path)

        # Check if the document contains paragraphs
        if len(doc.paragraphs) > 0:
            return True, "Valid document file."
        else:
            print(f"Document {file_path} is empty.")
            return False, f"Document {file_path} is empty."
        
    except Exception as e:
        print(f"Document {file_path} is not valid.")
        return False, f"Document {file_path} is not valid."

def validate_excel(file_path):
    from openpyxl import load_workbook

    try:
        # Try to open the workbook
        workbook = load_workbook(file_path)
        # Check if the workbook contains any sheets
        if not workbook.sheetnames:
            print(f"Invalid .xlsx file: {file_path} (No sheets found).")
            return False, f"Invalid .xlsx file: {file_path} (No sheets found)."
        else:
            return True, "Valid .xlsx file."
    except Exception as e:
        print(f"Invalid .xlsx file: {file_path} (Error: {e}).")
        return False, f"Invalid .xlsx file: {file_path} (Error: {e})."


def validate_pdf(file_path):
    try:
        # Importing PyPDF2 inside the function
        from PyPDF2 import PdfReader

        # Open and validate the PDF
        reader = PdfReader(file_path)
        if reader.pages:
            return True, "Valid pdf file."
        else:
            print(f"Invalid PDF: {file_path} (No pages found)")
            return False, f"Invalid PDF: {file_path} (No pages found)"

    except Exception as e:
        print(f"Invalid PDF: {file_path} (Error: {e})")
        return False, f"Invalid PDF: {file_path} (Error: {e})"


def validate_video(file_path):
    import subprocess

    try:
        # Run FFprobe command
        command = [
            "ffprobe",
            "-v", "error",  # Only show errors
            "-show_entries", "format=duration",  # Check for duration (validity indicator)
            "-of", "default=noprint_wrappers=1:nokey=1",
            file_path
        ]
        subprocess.check_output(command, stderr=subprocess.STDOUT)
        return True, "Valid video file." 
    except subprocess.CalledProcessError as e:
        # If FFprobe returns an error, the file is likely corrupted
        print (f"Video file {file_path} is not a valid video file.")
        return False, f"Video file {file_path} is not a valid video file."

def get_arguments(argv):
    arg_help = "{0} <directory> <filetype>".format(argv[0])

    try:
        arg_directory = (
            sys.argv[1]
            if len(sys.argv) > 1
            else '"C:\\Users\\vszal\\OneDrive\\Pictures"'
        )
        arg_filetypes = sys.argv[2] if len(sys.argv) > 2 else "images"
    except:
        print(arg_help)
        sys.exit(2)

    print("directory:", arg_directory)
    print("filetypes:", arg_filetypes)

    return [arg_directory, arg_filetypes]


def validate_image(file_path):
    try:
        with Image.open(file_path) as img:
            img.verify()  # Verify that it is a valid image
        return True, "Valid image file." # the file is not invalid
    except Exception as e:
        print(f"Invalid image file: {e}")
        return False, f"Invalid image file: {e}"

def get_total_file_count(directory):
    file_count = 0

    for root, dirs, files in os.walk(directory):
        file_count += len(files)

    return file_count

def count_files_with_pattern(directory, compiled_pattern):
    # Compile the regex pattern
    #compiled_pattern = re.compile(pattern)
    file_count = 0

    # Walk through the directory and subdirectories
    for root, dirs, files in os.walk(directory):
        for file in files:
            if compiled_pattern.match(file):  # Check if file matches the pattern
                file_count += 1

    return file_count

if __name__ == "__main__":
    arguments = get_arguments(sys.argv)
    print("directory: ", arguments[0])
    print("filetypes: ", arguments[1])

    # List files
    list_files_by_regex(arguments[0], arguments[1])

    # Print results
    # print(f"Found {len(matching_files)} matching files:")
    print("Analysis complete.")


# file_path = r"C:\Users\vszal\Downloads\IMG_2724.jpg"  # Use raw string for Windows paths

# print(validate_image_with_pillow(file_path))


# file_path = r"C:\Users\vszal\OneDrive\Pictures\5_27_22, 4_46 PM Microsoft Lens.jpg"
# print(validate_image_with_pillow(file_path))

# file_path = r"C:\Users\vszal\OneDrive\Pictures\The Sequel.pdf"
# print(validate_image_with_pillow(file_path))
